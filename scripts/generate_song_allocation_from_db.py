from __future__ import annotations

import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "db" / "sqm_inventory.db"
OUT_DIR = ROOT / "output"

ALLOC_DATE = date(2026, 7, 11)
CUSTOMER = "SONG SIMULATION"
SALE_REF = f"SONG-ALLOC-{ALLOC_DATE:%Y%m%d}"
STORAGE_FALLBACK = "GY"


def _columns(con: sqlite3.Connection, table: str) -> set[str]:
    return {row[1] for row in con.execute(f"PRAGMA table_info({table})").fetchall()}


def _weight_expr(cols: set[str]) -> str:
    if "weight_kg" in cols and "weight" in cols:
        return "COALESCE(t.weight_kg, t.weight, 0)"
    if "weight_kg" in cols:
        return "COALESCE(t.weight_kg, 0)"
    if "weight" in cols:
        return "COALESCE(t.weight, 0)"
    return "0"


def load_lots() -> list[sqlite3.Row]:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"DB not found: {DB_PATH}")

    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    try:
        tb_cols = _columns(con, "inventory_tonbag")
        inv_cols = _columns(con, "inventory")
        weight = _weight_expr(tb_cols)
        sample_expr = "(COALESCE(t.is_sample, 0)=1 OR COALESCE(t.sub_lt, 0)=0)"

        select_bits = [
            "t.lot_no",
            "COUNT(CASE WHEN NOT " + sample_expr + " THEN 1 END) AS normal_count",
            "SUM(CASE WHEN NOT " + sample_expr + " THEN " + weight + " ELSE 0 END) AS normal_kg",
            "COUNT(CASE WHEN " + sample_expr + " THEN 1 END) AS sample_count",
            "SUM(CASE WHEN " + sample_expr + " THEN " + weight + " ELSE 0 END) AS sample_kg",
        ]
        for col in ("sap_no", "product", "warehouse", "customs", "stock_date"):
            if col in inv_cols:
                select_bits.append(f"MAX(COALESCE(i.{col}, '')) AS {col}")
            else:
                select_bits.append(f"'' AS {col}")

        rows = con.execute(
            "SELECT "
            + ", ".join(select_bits)
            + """
            FROM inventory_tonbag t
            LEFT JOIN inventory i ON i.lot_no = t.lot_no
            WHERE NULLIF(TRIM(COALESCE(t.lot_no, '')), '') IS NOT NULL
            GROUP BY t.lot_no
            ORDER BY t.lot_no
            """
        ).fetchall()
        return [
            row
            for row in rows
            if int(row["normal_count"] or 0) > 0 or int(row["sample_count"] or 0) > 0
        ]
    finally:
        con.close()


def _product_name(row: sqlite3.Row, sample: bool = False) -> str:
    product = str(row["product"] or "LITHIUM CARBONATE").strip() or "LITHIUM CARBONATE"
    if sample and not product.upper().endswith(" SP"):
        return f"{product} SP"
    return product


def build_workbook(rows: list[sqlite3.Row]) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary = wb.active
    summary.title = "Sheet1"
    data = wb.create_sheet("250톤 수출작업")

    normal_count = sum(int(r["normal_count"] or 0) for r in rows)
    total_sample_count = sum(int(r["sample_count"] or 0) for r in rows)
    normal_mt = sum(float(r["normal_kg"] or 0) for r in rows) / 1000.0
    sample_mt = sum(float(r["sample_kg"] or 0) for r in rows) / 1000.0

    title = f"Allocation - {CUSTOMER} - July 2026 / CIF Busan - Song format"
    summary["A1"] = title
    summary["A1"].font = Font(bold=True, size=14)
    summary.append([])
    summary.append(["Outbound Date", ALLOC_DATE.isoformat()])
    summary.append(["Sale Ref", SALE_REF])
    summary.append(["LOT Count", len(rows)])
    summary.append(["Normal Tonbags", normal_count])
    summary.append(["SP Count", total_sample_count])
    summary.append(["Normal Qty (MT)", round(normal_mt, 4)])
    summary.append(["SP Qty (MT)", round(sample_mt, 4)])
    summary.append(["Total Qty (MT)", round(normal_mt + sample_mt, 4)])
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 24

    data["A1"] = title
    data["A1"].font = Font(bold=True, size=14)
    data["A2"] = round(normal_mt + sample_mt, 4)
    data["B2"] = "TOTAL QTY (MT)"

    headers = [
        "Product",
        "SAP NO",
        "Date in stock",
        "QTY (MT)",
        "Lot No",
        "WH",
        "Customs",
        "SOLD TO",
        "SALE REF",
        "OUTBOUND_DATE",
        "GW",
        "Remark",
    ]
    for col, header in enumerate(headers, start=1):
        cell = data.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cell.alignment = Alignment(horizontal="center")

    sample_fill = PatternFill("solid", fgColor="FFF2CC")
    row_idx = 4
    for item in rows:
        lot_no = str(item["lot_no"]).strip()
        sap_no = str(item["sap_no"] or "").strip()
        stock_date = str(item["stock_date"] or "").strip()
        warehouse = str(item["warehouse"] or STORAGE_FALLBACK).strip() or STORAGE_FALLBACK
        customs = str(item["customs"] or "").strip()

        normal_kg = float(item["normal_kg"] or 0)
        normal_qty = round(normal_kg / 1000.0, 4)
        if normal_qty > 0:
            values = [
                _product_name(item),
                sap_no,
                stock_date,
                normal_qty,
                lot_no,
                warehouse,
                customs,
                CUSTOMER,
                SALE_REF,
                ALLOC_DATE.isoformat(),
                round(normal_kg, 3),
                f"ALL NORMAL TONBAGS ALLOCATED ({int(item['normal_count'] or 0)} EA)",
            ]
            for col, value in enumerate(values, start=1):
                data.cell(row=row_idx, column=col, value=value)
            row_idx += 1

        lot_sample_count = int(item["sample_count"] or 0)
        sample_kg = float(item["sample_kg"] or 0)
        if lot_sample_count > 0:
            sample_qty = round((sample_kg if sample_kg > 0 else lot_sample_count) / 1000.0, 4)
            values = [
                _product_name(item, sample=True),
                sap_no,
                stock_date,
                sample_qty,
                lot_no,
                warehouse,
                customs,
                CUSTOMER,
                SALE_REF,
                ALLOC_DATE.isoformat(),
                round(sample_kg if sample_kg > 0 else sample_count, 3),
                f"SP ALLOCATED ({lot_sample_count} EA)",
            ]
            for col, value in enumerate(values, start=1):
                cell = data.cell(row=row_idx, column=col, value=value)
                cell.fill = sample_fill
            row_idx += 1

    widths = [28, 14, 14, 12, 16, 10, 14, 20, 22, 16, 12, 34]
    for idx, width in enumerate(widths, start=1):
        data.column_dimensions[get_column_letter(idx)].width = width
    data.freeze_panes = "A4"

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUT_DIR / f"song_allocation_from_db_20260711_{stamp}.xlsx"
    wb.save(out_path)
    print(f"file={out_path}")
    print(f"lots={len(rows)}")
    print(f"normal_tonbags={normal_count}")
    print(f"sp_count={total_sample_count}")
    print(f"normal_mt={round(normal_mt, 4)}")
    print(f"sp_mt={round(sample_mt, 4)}")
    print(f"sale_ref={SALE_REF}")
    print(f"outbound_date={ALLOC_DATE.isoformat()}")
    return out_path


def main() -> None:
    rows = load_lots()
    if not rows:
        raise SystemExit("No inventory_tonbag rows found.")
    build_workbook(rows)


if __name__ == "__main__":
    main()
