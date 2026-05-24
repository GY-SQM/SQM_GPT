from __future__ import annotations

import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "db" / "sqm_inventory.db"
OUT_DIR = ROOT / "output"

ALLOC_DATE = date(2026, 7, 11)
TITLE = "SQM Allocation Test - Song/AAA 10Col (Real LOTs)"
CUSTOMERS = ["LGES KOREA", "AAA BATTERY", "CATL KOREA", "LGES CHINA"]
WH_FALLBACK = "GY-LOGIS"
CUSTOMS_FALLBACK = "BONDED"


def load_rows() -> list[sqlite3.Row]:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    try:
        return con.execute(
            """
            SELECT
                i.product,
                i.sap_no,
                i.stock_date,
                i.inbound_date,
                i.warehouse,
                i.customs,
                i.sold_to,
                i.sale_ref,
                i.lot_no,
                i.current_weight,
                i.net_weight,
                i.gross_weight,
                COALESCE(SUM(CASE WHEN COALESCE(t.is_sample, 0)=0 AND COALESCE(t.sub_lt, 0)<>0 THEN t.weight ELSE 0 END), 0) AS normal_weight_kg,
                COALESCE(SUM(CASE WHEN COALESCE(t.is_sample, 0)=1 OR COALESCE(t.sub_lt, 0)=0 THEN t.weight ELSE 0 END), 0) AS sp_weight_kg,
                COUNT(CASE WHEN COALESCE(t.is_sample, 0)=0 AND COALESCE(t.sub_lt, 0)<>0 THEN 1 END) AS normal_count,
                COUNT(CASE WHEN COALESCE(t.is_sample, 0)=1 OR COALESCE(t.sub_lt, 0)=0 THEN 1 END) AS sp_count
            FROM inventory i
            LEFT JOIN inventory_tonbag t ON t.lot_no = i.lot_no
            WHERE NULLIF(TRIM(COALESCE(i.lot_no, '')), '') IS NOT NULL
            GROUP BY
                i.product, i.sap_no, i.stock_date, i.inbound_date, i.warehouse,
                i.customs, i.sold_to, i.sale_ref, i.lot_no,
                i.current_weight, i.net_weight, i.gross_weight
            ORDER BY i.lot_no
            """
        ).fetchall()
    finally:
        con.close()


def first_value(*values, default=""):
    for value in values:
        if value is not None and str(value).strip():
            return value
    return default


def product_name(row: sqlite3.Row) -> str:
    product = str(first_value(row["product"], default="LITHIUM CARBONATE")).strip()
    return product


def lot_qty_mt(row: sqlite3.Row) -> float:
    normal_kg = float(row["normal_weight_kg"] or 0)
    sp_kg = float(row["sp_weight_kg"] or 0)
    if normal_kg > 0 or sp_kg > 0:
        return round((normal_kg + sp_kg) / 1000.0, 4)
    current_kg = float(row["current_weight"] or 0)
    return round(current_kg / 1000.0, 4)


def gross_mt(row: sqlite3.Row, qty_mt: float) -> float:
    gross_kg = float(row["gross_weight"] or 0)
    if gross_kg > 0:
        return round(gross_kg / 1000.0, 4)
    return round(qty_mt * 1.02, 4)


def build() -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = load_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Allocation"

    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True, size=13)

    headers = [
        "Product",
        "SAP NO",
        "Date in stock",
        "QTY(MT)",
        "Lot No",
        "WH",
        "Customs",
        "SOLD TO",
        "SALE REF",
        "GW",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cell.alignment = Alignment(horizontal="center")

    normal_tonbag_count = 0
    sp_count = 0
    total_mt = 0.0

    for excel_row, row in enumerate(rows, start=3):
        idx = excel_row - 3
        qty_mt = lot_qty_mt(row)
        total_mt += qty_mt
        normal_tonbag_count += int(row["normal_count"] or 0)
        sp_count += int(row["sp_count"] or 0)

        values = [
            product_name(row),
            first_value(row["sap_no"]),
            ALLOC_DATE.isoformat(),
            qty_mt,
            row["lot_no"],
            first_value(row["warehouse"], default=WH_FALLBACK),
            first_value(row["customs"], default=CUSTOMS_FALLBACK),
            first_value(row["sold_to"], default=CUSTOMERS[idx % len(CUSTOMERS)]),
            first_value(row["sale_ref"], default=f"SC-2026-{idx + 1:03d}"),
            gross_mt(row, qty_mt),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=excel_row, column=col, value=value)

    widths = [44, 14, 14, 12, 16, 12, 14, 18, 22, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A3"

    out_path = OUT_DIR / f"song_allocation_lot_level_db_20260711_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(out_path)
    print(f"file={out_path}")
    print("sheets=1")
    print(f"rows={len(rows)}")
    print(f"normal_tonbags={normal_tonbag_count}")
    print(f"sp_rows={sp_count}")
    print(f"total_mt={round(total_mt, 4)}")
    return out_path


if __name__ == "__main__":
    build()
