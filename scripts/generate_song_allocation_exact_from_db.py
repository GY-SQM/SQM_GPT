from __future__ import annotations

import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "db" / "sqm_inventory.db"
OUT_DIR = ROOT / "output"

ALLOC_DATE = date(2026, 7, 11)
SALE_REF = f"SONG-ALLOC-{ALLOC_DATE:%Y%m%d}"
CUSTOMER_FALLBACK = "SONG"
WH_FALLBACK = "GY"


def connect() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def table_rows(con: sqlite3.Connection, table: str) -> tuple[list[str], list[sqlite3.Row]]:
    cols = [r[1] for r in con.execute(f"PRAGMA table_info({table})").fetchall()]
    rows = con.execute(f"SELECT * FROM {table} ORDER BY id").fetchall()
    return cols, rows


def load_tonbags(con: sqlite3.Connection) -> list[sqlite3.Row]:
    return con.execute(
        """
        SELECT
            t.id AS tonbag_id,
            t.inventory_id,
            t.lot_no,
            t.sap_no AS tonbag_sap_no,
            i.sap_no AS inventory_sap_no,
            t.bl_no AS tonbag_bl_no,
            i.bl_no AS inventory_bl_no,
            i.container_no,
            t.inbound_date AS tonbag_inbound_date,
            i.stock_date,
            i.inbound_date AS inventory_inbound_date,
            t.sub_lt,
            t.weight,
            t.is_sample,
            t.status AS tonbag_status,
            t.location AS tonbag_location,
            t.picked_to,
            t.picked_date,
            t.pick_ref,
            t.outbound_date AS tonbag_outbound_date,
            t.sale_ref AS tonbag_sale_ref,
            t.tonbag_uid,
            t.source_sub_lt_raw,
            t.source_sub_lt_hdr,
            t.con_return AS tonbag_con_return,
            t.tonbag_no,
            t.remarks AS tonbag_remarks,
            i.product,
            i.product_code,
            i.warehouse,
            i.customs,
            i.sold_to,
            i.sale_ref AS inventory_sale_ref,
            i.vessel,
            i.voyage,
            i.do_no,
            i.folio,
            i.status AS lot_status
        FROM inventory_tonbag t
        LEFT JOIN inventory i ON i.lot_no = t.lot_no
        WHERE NULLIF(TRIM(COALESCE(t.lot_no, '')), '') IS NOT NULL
        ORDER BY t.lot_no, COALESCE(t.is_sample, 0), t.sub_lt, t.id
        """
    ).fetchall()


def append_table(ws, headers: Iterable[str], rows: Iterable[sqlite3.Row], start_row: int = 1) -> None:
    headers = list(headers)
    for _ in range(max(0, start_row - 1)):
        ws.append([])
    ws.append(headers)
    for cell in ws[start_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row[h] for h in headers])
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 2, 10), 34)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = f"A{start_row + 1}"


def _value(row: sqlite3.Row, *names: str, default: str = ""):
    for name in names:
        val = row[name]
        if val is not None and str(val).strip() != "":
            return val
    return default


def _product(row: sqlite3.Row, is_sample: bool) -> str:
    product = str(_value(row, "product", default="")).strip()
    if is_sample and product and not product.upper().endswith(" SP"):
        return f"{product} SP"
    return product


def build_workbook() -> Path:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"DB not found: {DB_PATH}")
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    with connect() as con:
        tonbags = load_tonbags(con)
        inv_cols, inv_rows = table_rows(con, "inventory")
        tb_cols, tb_rows = table_rows(con, "inventory_tonbag")

    wb = Workbook()
    summary = wb.active
    summary.title = "Sheet1"
    data = wb.create_sheet("250톤 수출작업")
    raw_inv = wb.create_sheet("DB_inventory")
    raw_tb = wb.create_sheet("DB_inventory_tonbag")

    normal_count = sum(1 for r in tonbags if not bool(r["is_sample"] or 0) and int(r["sub_lt"] or 0) != 0)
    sp_count = len(tonbags) - normal_count
    total_mt = round(sum(float(r["weight"] or 0) for r in tonbags) / 1000.0, 4)
    sp_mt = round(
        sum(float(r["weight"] or 0) for r in tonbags if bool(r["is_sample"] or 0) or int(r["sub_lt"] or 0) == 0)
        / 1000.0,
        4,
    )

    title = "Allocation - Song format - exact rows from DB inventory_tonbag"
    summary["A1"] = title
    summary["A1"].font = Font(bold=True, size=14)
    for row in [
        ("Outbound Date", ALLOC_DATE.isoformat()),
        ("Sale Ref", SALE_REF),
        ("DB Source", str(DB_PATH)),
        ("Inventory LOT Rows", len(inv_rows)),
        ("Inventory Tonbag Rows", len(tb_rows)),
        ("Song Data Rows", len(tonbags)),
        ("Normal Tonbags", normal_count),
        ("SP Rows", sp_count),
        ("Total Qty (MT)", total_mt),
        ("SP Qty (MT)", sp_mt),
    ]:
        summary.append(row)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 80

    data["A1"] = title
    data["A1"].font = Font(bold=True, size=14)
    data["A2"] = total_mt
    data["B2"] = "TOTAL QTY (MT)"

    headers = [
        "Product",
        "SAP NO",
        "Date in stock",
        "QTY (MT)",
        "Lot No",
        "Tonbag No",
        "WH",
        "Customs",
        "SOLD TO",
        "SALE REF",
        "OUTBOUND_DATE",
        "GW",
        "Tonbag ID",
        "Tonbag UID",
        "Status",
        "Location",
        "BL NO",
        "Container No",
        "Remark",
    ]
    for col, header in enumerate(headers, start=1):
        cell = data.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cell.alignment = Alignment(horizontal="center")

    sp_fill = PatternFill("solid", fgColor="FFF2CC")
    row_idx = 4
    for item in tonbags:
        is_sp = bool(item["is_sample"] or 0) or int(item["sub_lt"] or 0) == 0
        weight_kg = float(item["weight"] or 0)
        values = [
            _product(item, is_sp),
            _value(item, "tonbag_sap_no", "inventory_sap_no"),
            _value(item, "stock_date", "tonbag_inbound_date", "inventory_inbound_date"),
            round(weight_kg / 1000.0, 4),
            item["lot_no"],
            item["sub_lt"],
            _value(item, "warehouse", default=WH_FALLBACK),
            _value(item, "customs"),
            _value(item, "sold_to", default=CUSTOMER_FALLBACK),
            _value(item, "tonbag_sale_ref", "inventory_sale_ref", default=SALE_REF),
            ALLOC_DATE.isoformat(),
            round(weight_kg, 3),
            item["tonbag_id"],
            item["tonbag_uid"],
            item["tonbag_status"],
            item["tonbag_location"],
            _value(item, "tonbag_bl_no", "inventory_bl_no"),
            item["container_no"],
            _value(item, "tonbag_remarks"),
        ]
        for col, value in enumerate(values, start=1):
            cell = data.cell(row=row_idx, column=col, value=value)
            if is_sp:
                cell.fill = sp_fill
        row_idx += 1

    widths = [28, 14, 14, 12, 16, 10, 10, 14, 20, 22, 16, 10, 12, 24, 12, 18, 16, 18, 28]
    for idx, width in enumerate(widths, start=1):
        data.column_dimensions[get_column_letter(idx)].width = width
    data.freeze_panes = "A4"

    raw_inv["A1"] = "Raw DB table copy: inventory"
    raw_inv["A1"].font = Font(bold=True)
    raw_tb["A1"] = "Raw DB table copy: inventory_tonbag"
    raw_tb["A1"].font = Font(bold=True)
    append_table(raw_inv, inv_cols, inv_rows, start_row=12)
    append_table(raw_tb, tb_cols, tb_rows, start_row=12)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUT_DIR / f"song_allocation_exact_db_rows_20260711_{stamp}.xlsx"
    wb.save(out_path)
    print(f"file={out_path}")
    print(f"inventory_lots={len(inv_rows)}")
    print(f"inventory_tonbags={len(tb_rows)}")
    print(f"song_rows={len(tonbags)}")
    print(f"normal_tonbags={normal_count}")
    print(f"sp_rows={sp_count}")
    print(f"total_mt={total_mt}")
    print(f"sp_mt={sp_mt}")
    print(f"sale_ref={SALE_REF}")
    print(f"outbound_date={ALLOC_DATE.isoformat()}")
    return out_path


if __name__ == "__main__":
    build_workbook()
