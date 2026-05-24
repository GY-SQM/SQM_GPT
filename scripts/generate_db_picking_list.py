# -*- coding: utf-8 -*-
"""Generate a simulation Picking List Excel from the current SQM DB.

This does not modify the database. It treats all current tonbags as eligible
for picking-list simulation, which matches the "assume all resolved" workflow.
"""
from __future__ import annotations

import os
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "db" / "sqm_inventory.db"
OUT_DIR = ROOT / "output"
STORAGE_FALLBACK = "1001 GY logistics"


def _columns(con: sqlite3.Connection, table: str) -> set[str]:
    return {str(r[1]) for r in con.execute(f"PRAGMA table_info({table})")}


def _weight_expr(tb_cols: set[str]) -> str:
    if "weight_kg" in tb_cols and "weight" in tb_cols:
        return "COALESCE(t.weight_kg, t.weight, 0)"
    if "weight_kg" in tb_cols:
        return "COALESCE(t.weight_kg, 0)"
    if "weight" in tb_cols:
        return "COALESCE(t.weight, 0)"
    return "0"


def load_lot_rows() -> list[dict]:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    try:
        tb_cols = _columns(con, "inventory_tonbag")
        inv_cols = _columns(con, "inventory")
        weight = _weight_expr(tb_cols)

        location_bits = []
        if "location" in tb_cols:
            location_bits.append("NULLIF(TRIM(t.location), '')")
        if "location" in inv_cols:
            location_bits.append("NULLIF(TRIM(i.location), '')")
        if "warehouse" in inv_cols:
            location_bits.append("NULLIF(TRIM(i.warehouse), '')")
        loc_expr = "COALESCE(" + ", ".join(location_bits + [f"'{STORAGE_FALLBACK}'"]) + ")"

        sample_expr = "(COALESCE(t.is_sample, 0)=1 OR COALESCE(t.sub_lt, 0)=0)"
        sql = f"""
            SELECT
                t.lot_no,
                COUNT(CASE WHEN NOT {sample_expr} THEN 1 END) AS normal_count,
                SUM(CASE WHEN NOT {sample_expr} THEN {weight} ELSE 0 END) AS normal_kg,
                COUNT(CASE WHEN {sample_expr} THEN 1 END) AS sample_count,
                SUM(CASE WHEN {sample_expr} THEN {weight} ELSE 0 END) AS sample_kg,
                MIN({loc_expr}) AS storage_location,
                MAX(COALESCE(i.product, '')) AS product
            FROM inventory_tonbag t
            LEFT JOIN inventory i ON i.lot_no = t.lot_no
            WHERE NULLIF(TRIM(COALESCE(t.lot_no, '')), '') IS NOT NULL
            GROUP BY t.lot_no
            ORDER BY t.lot_no
        """
        rows = [dict(r) for r in con.execute(sql)]
        return [r for r in rows if int(r.get("normal_count") or 0) or int(r.get("sample_count") or 0)]
    finally:
        con.close()


def write_excel(rows: list[dict]) -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUT_DIR / f"picking_list_sim_from_db_{stamp}.xlsx"

    creation = datetime.now().strftime("%Y-%m-%d")
    loading = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
    picking_no = f"PK-SIM-DB-{stamp}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PickingList"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="C00000")
    sample_fill = PatternFill("solid", fgColor="FFF2CC")
    center = Alignment(horizontal="center")

    ws["A1"] = "PICKING LIST"
    ws["A1"].font = title_font
    kv = [
        ("Outbound ID", f"OUT-SIM-DB-{stamp}"),
        ("Sales Order", f"SO-SIM-DB-{stamp}"),
        ("Customer reference", picking_no),
        ("Customer", "HANDS DB SIMULATION"),
        ("Creation Date", creation),
        ("Plan Loading Date", loading),
    ]
    for idx, (key, value) in enumerate(kv, start=2):
        ws.cell(row=idx, column=1, value=key).font = bold
        ws.cell(row=idx, column=2, value=value)

    hdr_row = 9
    headers = ["Lot No", "Quantity", "Unit", "Storage location"]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=col, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center

    excel_row = hdr_row + 1
    normal_total = 0
    sample_total = 0
    normal_mt_total = 0.0
    sample_kg_total = 0.0

    for item in rows:
        lot_no = str(item["lot_no"])
        storage = str(item.get("storage_location") or STORAGE_FALLBACK)
        normal_count = int(item.get("normal_count") or 0)
        sample_count = int(item.get("sample_count") or 0)
        normal_kg = float(item.get("normal_kg") or 0)
        sample_kg = float(item.get("sample_kg") or 0)

        if normal_count:
            qty_mt = round(normal_kg / 1000.0, 4)
            ws.cell(row=excel_row, column=1, value=lot_no)
            ws.cell(row=excel_row, column=2, value=qty_mt)
            ws.cell(row=excel_row, column=3, value="MT")
            ws.cell(row=excel_row, column=4, value=storage)
            excel_row += 1
            normal_total += normal_count
            normal_mt_total += qty_mt

        if sample_count:
            qty_kg = round(sample_kg if sample_kg > 0 else float(sample_count), 4)
            for col, value in enumerate([lot_no, qty_kg, "KG", storage], start=1):
                cell = ws.cell(row=excel_row, column=col, value=value)
                cell.fill = sample_fill
            excel_row += 1
            sample_total += sample_count
            sample_kg_total += qty_kg

    for col, width in enumerate([18, 12, 8, 28], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    summary_row = excel_row + 1
    ws.cell(row=summary_row, column=1, value="Summary").font = bold
    ws.cell(row=summary_row + 1, column=1, value="LOT count")
    ws.cell(row=summary_row + 1, column=2, value=len(rows))
    ws.cell(row=summary_row + 2, column=1, value="Normal tonbags")
    ws.cell(row=summary_row + 2, column=2, value=normal_total)
    ws.cell(row=summary_row + 3, column=1, value="Sample bags")
    ws.cell(row=summary_row + 3, column=2, value=sample_total)
    ws.cell(row=summary_row + 4, column=1, value="Normal total MT")
    ws.cell(row=summary_row + 4, column=2, value=round(normal_mt_total, 4))
    ws.cell(row=summary_row + 5, column=1, value="Sample total KG")
    ws.cell(row=summary_row + 5, column=2, value=round(sample_kg_total, 4))

    wb.save(out_path)
    print(f"file={out_path}")
    print(f"lots={len(rows)}")
    print(f"normal_tonbags={normal_total}")
    print(f"samples={sample_total}")
    print(f"normal_mt={round(normal_mt_total, 4)}")
    print(f"sample_kg={round(sample_kg_total, 4)}")
    print(f"picking_no={picking_no}")
    return out_path


def main() -> None:
    rows = load_lot_rows()
    if not rows:
        raise SystemExit("No tonbag rows found in DB.")
    write_excel(rows)


if __name__ == "__main__":
    main()
